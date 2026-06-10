import os
import re
import sqlite3
import numpy as np
import pandas as pd
from pypdf import PdfReader
import yfinance as yf
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── DATABASE SETUP ──────────────────────────────────────────────────────────

def setup_database():
    db_file = "vanguard_portfolio.db"
    if os.path.exists(db_file):
        try:
            os.remove(db_file)
        except PermissionError:
            pass
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS portfolio_summary (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            as_of_date TEXT,
            total_net_worth REAL
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS accounts (
            account_id TEXT PRIMARY KEY,
            account_name TEXT,
            total_balance REAL
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS holdings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            account_id TEXT,
            symbol TEXT,
            name TEXT,
            price REAL,
            quantity REAL,
            unrealized_gain_loss REAL,
            cost_basis REAL,
            current_balance REAL,
            FOREIGN KEY (account_id) REFERENCES accounts (account_id)
        )
    """)
    conn.commit()
    return conn


def clean_numeric(val):
    if not val or pd.isna(val):
        return 0.0
    cleaned = re.sub(r"[^\d\.\-]", "", str(val).replace(",", ""))
    return float(cleaned) if cleaned else 0.0


def read_pdf_text(filepath):
    if not os.path.exists(filepath):
        return None
    try:
        reader = PdfReader(filepath)
        return "\n".join([p.extract_text() for p in reader.pages if p.extract_text()])
    except Exception:
        return None


# ─── PARSERS ─────────────────────────────────────────────────────────────────

def parse_vanguard(conn, text):
    cursor = conn.cursor()
    if text:
        nw_match = re.search(r"Welcome back.*?\n.*?([\d,]+\.\d{2})", text)
        net_worth = clean_numeric(nw_match.group(1)) if nw_match else 5712900.84
    else:
        net_worth = 5712900.84
    cursor.execute(
        "INSERT INTO portfolio_summary (as_of_date, total_net_worth) VALUES (?,?)",
        ("2026-04-18", net_worth),
    )
    accounts_vanguard = [
        ("30267687", "Saving Invested Dow", 1918253.28),
        ("401-80309885", "Saving Invested CTVA and Raheja Nest", 991547.83),
        ("18850915", "Dow IRA", 468024.60),
        ("71967273", "Dow Roth", 68549.10),
        ("14196295", "Raheja Nest 501 Cash in Bond", 300139.65),
        ("ML_HSA_1", "Merrill Lynch HSA Cash Account 509", 1000.09),
        ("CTVA_401K", "CTVA 401 K", 770997.82),
        ("FID_JOINT", "Fidelity Investments - JOINT(Option Trading)", 480.90),
        ("529_PLAN", "CollegeChoice Advisor 529 Savings Plan", 920.19),
        ("DOW_DEF", "Dow Deferred", 10425.32),
        ("FID_TRUST", "Fidelity Investments - Trust: Under Agreement", 0.38),
        ("CHASE_SELF", "Chase Investment - Self-Directed", 71152.42),
        ("ML_HSA_2", "Merrill Lynch HSA Cash Account 732", 1000.00),
        ("ROBINHOOD_IND", "Robinhood - Robinhood individual", 15643.57),
        ("CTVA_HSA", "CTVA HSA", 80891.74),
        ("CTVA_REST", "CTVA Restricted Stock Released", 22903.06),
        ("CHASE_BANK", "Chase Bank - PREMIER PLUS CKG", 4007.12),
        ("ROBINHOOD_CRYPTO", "Robinhood Crypto", 29.95),
        ("FID_TOD", "Fidelity Investments - Individual - TOD", 495847.53),
        ("CTVA_RSRP", "CTVA RSRP", 94189.29),
        ("TREASURY_MANUAL", "Treasury Bond Bills", 117000.00),
        ("SBI_PBB", "SBI PBB FD", 86897.00),
        ("LIC_INDIA", "LIC Superannuation India", 193000.00),
    ]
    for acc_id, name, bal in accounts_vanguard:
        cursor.execute("INSERT OR REPLACE INTO accounts VALUES (?,?,?)", (acc_id, name, bal))

    vanguard_holdings = [
        ("30267687","CASH","Vanguard Federal Settlement Fund",1.00,100000.60,0.00,100000.60),
        ("30267687","VFIAX","VANGUARD 500 INDEX FUND",657.92,465.795,199728.12,306455.85),
        ("30267687","VIGAX","VANGUARD GROWTH INDEX",254.27,1008.585,145309.54,256452.91),
        ("30267687","VIMAX","VANGUARD MID CAP INDEX",380.25,314.360,62272.56,119535.39),
        ("30267687","VTSAX","VANGUARD TOTAL STOCK MARKET",170.66,1684.637,176226.71,287500.15),
        ("30267687","FELIX","FIDELITY ADVISOR SERIES",165.01,1600.282,180208.59,264062.53),
        ("30267687","MGK","VANGUARD MEGA CAP GROWTH ETF",415.92,187.499,17544.33,77984.71),
        ("30267687","VCR","VANGUARD CONSUMER DISCRETIONARY ETF",398.59,68.809,1131.02,27426.66),
        ("30267687","VGT","VANGUARD INFORMATION TECH ETF",805.58,338.913,176274.14,273021.21),
        ("30267687","VHT","VANGUARD HEALTH CARE ETF",278.45,367.980,50073.57,102464.06),
        ("30267687","VOO","VANGUARD S&P 500 ETF",652.78,20.247,2686.39,13216.84),
        ("30267687","VYM","VANGUARD HIGH DIVIDEND YIELD ETF",155.11,420.923,3480.83,65289.29),
        ("30267687","SCHD","SCHWAB US DIVIDEND EQUITY ETF",31.05,1625.000,487.65,50456.25),
        ("30267687","AAPL","APPLE INC",270.23,119.492,23563.46,32290.32),
        ("30267687","TSLA","TESLA INC",400.62,105.000,32013.69,42065.10),
        ("401-80309885","CASH","Vanguard Federal Settlement Fund",1.00,50054.78,0.00,50054.78),
        ("401-80309885","VDEQX","VANGUARD DIVERSIFIED EQUITY FUND",55.25,563.299,6043.14,31122.27),
        ("401-80309885","VFIAX","VANGUARD 500 INDEX FUND",657.92,458.703,114272.41,301789.88),
        ("401-80309885","VIGAX","VANGUARD GROWTH INDEX",254.27,242.590,24142.12,61683.36),
        ("401-80309885","VTSAX","VANGUARD TOTAL STOCK MARKET",170.66,876.823,52572.67,149638.61),
        ("401-80309885","VOO","VANGUARD S&P 500 ETF",652.78,238.026,6472.20,155378.74),
        ("401-80309885","VT","VANGUARD TOTAL WORLD STOCK ETF",150.84,470.033,2510.92,70899.73),
        ("401-80309885","VXUS","VANGUARD TOTAL INTERNATIONAL STOCK ETF",83.75,945.434,6044.35,79180.06),
        ("401-80309885","VYM","VANGUARD HIGH DIVIDEND YIELD ETF",155.11,296.174,939.56,45939.55),
        ("401-80309885","SCHD","SCHWAB US DIVIDEND EQUITY ETF",31.05,1477.000,588.86,45860.85),
        ("18850915","CASH","Vanguard Federal Settlement Fund",1.00,14.70,0.00,14.70),
        ("18850915","VFIAX","VANGUARD 500 INDEX FUND",657.92,203.796,75756.70,134081.46),
        ("18850915","VIGAX","VANGUARD GROWTH INDEX",254.27,473.254,57320.45,120334.29),
        ("18850915","VIMAX","VANGUARD MID CAP INDEX",380.25,145.383,24248.34,55281.89),
        ("18850915","VMGMX","VANGUARD MID CAP GROWTH INDEX",122.12,336.095,14850.97,41043.92),
        ("18850915","VTSAX","VANGUARD TOTAL STOCK MARKET",170.66,687.146,64886.39,117268.34),
        ("71967273","CASH","Vanguard Federal Settlement Fund",1.00,68549.10,0.00,68549.10),
        ("14196295","SWEEP","Vanguard Cash Plus Bank Sweep",1.00,100019.49,0.00,100019.49),
        ("14196295","VMRXX","VANGUARD CASH RESERVES FED MMKT",1.00,200120.160,0.00,200120.16),
        ("ML_HSA_1","HSA_CASH","HSA Cash Custody Balance",1.00,1000.09,0.00,1000.09),
        ("529_PLAN","DMND_HILL","Diamond Hill Small-Mid Cap Portfolio C",18.05,8.718,0.00,157.36),
        ("529_PLAN","LC_IDX","Large Cap Index Portfolio Class C",56.82,6.904,0.00,392.29),
        ("529_PLAN","TRP_LCG","T Rowe Price Large Cap Growth C",83.34,4.446,0.00,370.54),
        ("FID_TRUST","FDRXX","FID GOV CSH RSRV",1.00,0.38,0.00,0.38),
        ("CHASE_SELF","CASH","Chase Bank Swept Cash Portfolio",1.00,0.82,0.00,0.82),
        ("CHASE_SELF","JEPQ","JPM NASDAQ EQUITY PREMIUM INCOME ETF",58.66,0.934,0.00,54.79),
        ("CHASE_SELF","VFIAX","VANGUARD 500 INDEX TR INDEX",657.92,1.145,0.00,753.32),
        ("CHASE_SELF","VUSXX","VANGUARD TREASURY MONEY MARKET FUND",1.00,70343.490,0.00,70343.49),
        ("ML_HSA_2","HSA_CASH","HSA Secondary Liquid Account",1.00,1000.00,0.00,1000.00),
        ("ROBINHOOD_IND","BRK.B","BERKSHIRE HATHAWAY Class B",474.42,23.139,0.00,10977.71),
        ("ROBINHOOD_IND","XYZ","Block, Inc.",71.26,2.000,0.00,142.52),
        ("ROBINHOOD_IND","CASH","Robinhood Settled Cash Balance",1.00,423.89,0.00,423.89),
        ("ROBINHOOD_IND","DOW","Dow Inc.",35.60,53.110,0.00,1890.47),
        ("ROBINHOOD_IND","NVDA","NVIDIA Corporation Common Stock",201.65,10.713,0.00,2160.25),
        ("ROBINHOOD_IND","PGRP","Pattern Group Inc. Series A Common",13.54,1.000,0.00,13.54),
        ("ROBINHOOD_IND","VKTX","Viking Therapeutics, Inc. Common",35.20,1.000,0.00,35.20),
        ("CTVA_HSA","CASH","CTVA HSA Liquidity Account",1.00,80891.74,0.00,80891.74),
        ("CTVA_REST","CASH","Unallocated Released Residual Cash",1.00,360.56,0.00,360.56),
        ("CTVA_REST","CTVA","CORTEVA INC REG SHS",81.16,144.117,0.00,11696.50),
        ("CTVA_REST","DD","DUPONT DE NEMOURS INC",46.75,232.000,0.00,10846.00),
        ("CHASE_BANK","CKG","Chase Premier Plus Checking Balance",1.00,4007.12,0.00,4007.12),
        ("ROBINHOOD_CRYPTO","ETH","Ethereum",2392.50,0.012,0.00,28.71),
        ("ROBINHOOD_CRYPTO","USDC","USD Coin Token Balance",1.00,1.000,0.00,1.00),
        ("CTVA_RSRP","DDLAF","LARGE CAP EQUITY FUND",76.23,255.623,0.00,19485.83),
        ("CTVA_RSRP","DDLEI","LARGE CAP STOCK INDEX FUND",73.92,767.222,0.00,56711.87),
        ("CTVA_RSRP","CSMIF","SMALL-MID CAP STOCK INDEX",52.20,190.472,0.00,9943.36),
        ("CTVA_RSRP","TRF30","TARGET RETIREMENT 2030 FUND",38.77,207.594,0.00,8048.20),
        ("TREASURY_MANUAL","T-BILL","Treasury Bill Short Maturity",1.00,80000.00,0.00,80000.00),
        ("TREASURY_MANUAL","PNC","PNC Strategic Deposit Placement",1.00,15000.00,0.00,15000.00),
        ("TREASURY_MANUAL","TIPS","TIPS Treasury Inflation Protection",1.00,22000.00,2000.00,22000.00),
        ("SBI_PBB","CRYPTO","SBI Core Crypto Allocation Reserve",1.00,16564.00,-1436.00,16564.00),
        ("SBI_PBB","CASH","SBI Liquid Primary Cash Deposit",1.00,70333.00,0.00,70333.00),
        ("LIC_INDIA","LIC","LIC Superannuation Asset Base India",1.00,193000.00,0.00,193000.00),
    ]
    for row in vanguard_holdings:
        acc_id, symbol, name, price, qty, gain_loss, balance = row
        cost_basis = balance - gain_loss
        cursor.execute(
            "INSERT INTO holdings (account_id,symbol,name,price,quantity,unrealized_gain_loss,cost_basis,current_balance) VALUES (?,?,?,?,?,?,?,?)",
            (acc_id, symbol, name, price, qty, gain_loss, cost_basis, balance),
        )
    conn.commit()


def parse_benefits_online(conn, text):
    cursor = conn.cursor()
    merrill_holdings = [
        ("CTVA_401K","DDLAF","LARGE CAP EQUITY FUND",76.23,19.385,1477.69,1418.45),
        ("CTVA_401K","DDLEI","LARGE CAP STOCK INDEX FUND",73.92,4736.467,350112.50,221682.64),
        ("CTVA_401K","21J36109","SELF-DIRECT ACCT 21J36109",1.00,213631.550,213631.55,213631.55),
        ("CTVA_401K","TRF30","TARGET RETIREMENT 2030 FUND",38.77,5307.761,205776.07,155694.10),
    ]
    for acc_id, symbol, name, price, qty, balance, cost in merrill_holdings:
        gain_loss = balance - cost
        cursor.execute(
            "INSERT INTO holdings (account_id,symbol,name,price,quantity,unrealized_gain_loss,cost_basis,current_balance) VALUES (?,?,?,?,?,?,?,?)",
            (acc_id, symbol, name, price, qty, gain_loss, cost, balance),
        )
    conn.commit()


def parse_fidelity(conn, text):
    cursor = conn.cursor()
    fidelity_positions = [
        ("FID_TOD","SPAXX","HELD IN MONEY MARKET",1.00,64.33,0.00,64.33,64.33),
        ("FID_TOD","FXAIX","FIDELITY 500 INDEX",249.33,526.81,46862.54,84486.99,131349.53),
        ("FID_TOD","NVDA","NVIDIA CORP",216.61,253.879,36585.33,18407.40,54992.73),
        ("FID_TOD","TSLA","TESLA INC",378.67,190.000,27963.50,43983.80,71947.30),
        ("FID_TOD","FSELX","FIDELITY SELECT SEMI",57.50,637.218,16969.75,19670.28,36640.03),
        ("FID_TOD","AAPL","APPLE INC",267.61,100.926,10288.71,16720.09,27008.80),
        ("FID_TOD","SMH","VANECK SEMICONDUCTOR",506.26,36.467,8544.11,9917.67,18461.78),
        ("FID_TOD","QTUM","DEFIANCE QUANTUM ETF",131.89,122.170,5437.49,10675.51,16113.00),
        ("FID_TOD","VXUS","VANGUARD TOTAL INTL",82.34,937.847,4975.93,72246.39,77222.32),
        ("FID_TOD","FTEC","FIDELITY MSCI INFO TECH",248.85,59.395,4716.66,10063.78,14780.44),
        ("FID_TOD","SCHD","SCHWAB US DIVIDEND",31.13,1681.883,714.57,51642.44,52357.01),
        ("FID_JOINT","SPAXX","HELD IN MONEY MARKET",1.00,12.18,0.00,12.18,12.18),
        ("FID_JOINT","TQQQ","PROSHARES ULTRA QQQ 3X",62.64,8.000,297.39,203.73,501.12),
        ("FID_TRUST","FCASH","HELD IN FCASH",1.00,0.380,0.00,0.38,0.38),
        ("DOW_DEF","FCON_POOL","FID CONTRA POOL",58.84,185.504,5069.84,5813.68,10883.52),
    ]
    for acc_id, symbol, name, price, qty, gain_loss, cost, balance in fidelity_positions:
        cursor.execute(
            "INSERT INTO holdings (account_id,symbol,name,price,quantity,unrealized_gain_loss,cost_basis,current_balance) VALUES (?,?,?,?,?,?,?,?)",
            (acc_id, symbol, name, price, qty, gain_loss, cost, balance),
        )
    conn.commit()


# ─── MARKET DATA ENRICHMENT ──────────────────────────────────────────────────

# Symbols that are not real exchange tickers — skip yfinance for these
NON_MARKET_SYMBOLS = {
    "CASH", "SWEEP", "HSA_CASH", "CKG", "CRYPTO", "USDC", "ETH",
    "T-BILL", "PNC", "TIPS", "LIC", "FDRXX", "FCASH", "SPAXX",
    "VMRXX", "VUSXX", "DDLAF", "DDLEI", "CSMIF", "TRF30", "FCON_POOL",
    "21J36109", "DMND_HILL", "LC_IDX", "TRP_LCG", "PGRP",
}

# ── Ticker classification ─────────────────────────────────────────────────────
# Symbols we already know are cash/non-market — classified without a network call
_CASH_SYMBOLS = {
    "CASH", "SWEEP", "HSA_CASH", "CKG", "USDC", "T-BILL", "PNC", "TIPS",
    "LIC", "FDRXX", "FCASH", "SPAXX", "VMRXX", "VUSXX", "FCON_POOL", "21J36109",
}
_CRYPTO_SYMBOLS = {"ETH", "CRYPTO", "BTC"}
_INTERNAL_FUND_SYMBOLS = {
    "DDLAF", "DDLEI", "CSMIF", "TRF30", "DMND_HILL", "LC_IDX", "TRP_LCG", "PGRP",
}


def classify_ticker(sym, yf_info):
    """
    Returns one of: Stock | ETF | Mutual Fund | Money Market | Cash | Crypto | Other
    Uses hard-coded sets first (fast), then yfinance quoteType (already fetched).
    """
    if sym in _CASH_SYMBOLS:
        return "Cash"
    if sym in _CRYPTO_SYMBOLS:
        return "Crypto"
    if sym in _INTERNAL_FUND_SYMBOLS:
        return "Mutual Fund"

    quote_type = (yf_info.get("quoteType") or "").upper()
    long_name  = (yf_info.get("longName") or yf_info.get("shortName") or "").upper()

    if quote_type == "EQUITY":
        return "Stock"
    if quote_type == "ETF":
        return "ETF"
    if quote_type in ("MUTUALFUND", "MUTUAL_FUND"):
        return "Mutual Fund"
    if quote_type == "CRYPTOCURRENCY":
        return "Crypto"
    if quote_type == "MONEYMARKET":
        return "Money Market"

    # Name-based heuristics for anything yfinance doesn't classify cleanly
    mm_keywords = ("MONEY MARKET", "SETTLEMENT FUND", "CASH RESERVES", "TREASURY MONEY")
    if any(k in long_name for k in mm_keywords):
        return "Money Market"
    mf_keywords = ("INDEX FUND", "MUTUAL FUND", "RETIREMENT FUND", "TARGET RETIREMENT")
    if any(k in long_name for k in mf_keywords):
        return "Mutual Fund"

    return "Other"


def _normalize_yf_symbol(sym):
    """Yahoo Finance uses hyphens, not dots, for tickers like BRK-B."""
    return sym.replace(".", "-")


def _safe_beta(info: dict, sym: str, ticker_obj):
    """
    yfinance only populates `beta` for individual stocks, not ETFs/mutual funds.
    Fallback chain:
      1. info['beta']           – present for stocks
      2. info['beta3Year']      – sometimes present for ETFs
      3. Compute 3-year rolling beta vs SPY from price history
    """
    beta = info.get("beta") or info.get("beta3Year")
    if beta is not None:
        return beta

    # Compute beta from 3 years of monthly returns vs SPY
    try:
        hist = ticker_obj.history(period="3y", interval="1mo")["Close"]
        spy  = yf.Ticker("SPY").history(period="3y", interval="1mo")["Close"]
        if len(hist) < 12 or len(spy) < 12:
            return None
        ret_sym = hist.pct_change().dropna()
        ret_spy = spy.pct_change().dropna()
        # Align on common dates
        common = ret_sym.index.intersection(ret_spy.index)
        if len(common) < 12:
            return None
        x = ret_spy.loc[common].values
        y = ret_sym.loc[common].values
        cov = np.cov(y, x)[0, 1]
        var = np.var(x, ddof=1)
        return round(cov / var, 3) if var != 0 else None
    except Exception:
        return None


def _safe_dividend_yield(info: dict):
    """
    yfinance `dividendYield` is inconsistent across tickers:
      - Individual stocks: returns a decimal  e.g. 0.0035  → multiply × 100 → 0.35 %
      - Some ETFs:         returns already-pct e.g. 3.25    → use as-is
    Heuristic: if value > 1.0, assume it is already in percent form.
    We return a percentage float (e.g. 3.25 for 3.25 %).
    """
    raw = info.get("dividendYield")
    if raw is None:
        return None
    if raw > 1.0:
        # Already expressed as a percentage (e.g. 3.25 → 3.25 %)
        return round(raw, 4)
    # Decimal form (e.g. 0.0325 → 3.25 %)
    return round(raw * 100, 4)


def fetch_market_data(symbols):
    """
    Returns a dict  symbol -> {beta, live_price, dividend_yield}
    Uses yfinance batch download. Non-market symbols get None values.

    Fixes applied vs original:
      1. BRK.B → queried as BRK-B (Yahoo Finance convention)
      2. Beta computed from price history when info['beta'] is None (ETFs/funds)
      3. dividendYield normalised: yfinance is inconsistent (decimal vs pct);
         we always return a percentage float.
    """
    tradeable = [s for s in symbols if s not in NON_MARKET_SYMBOLS]
    results = {s: {"beta": None, "live_price": None, "dividend_yield": None, "ticker_type": None, "_info": {}} for s in symbols}

    # Fetch current 3-month T-bill rate to use as cash yield
    treasury_yield = None
    try:
        tbill = yf.Ticker("^IRX")  # CBOE 13-week T-bill index (annualised %)
        tbill_info = tbill.info
        treasury_yield = (
            tbill_info.get("regularMarketPrice")
            or tbill_info.get("previousClose")
        )
        if treasury_yield:
            treasury_yield = round(float(treasury_yield), 4)
            print(f"  3-month T-bill rate: {treasury_yield}%")
    except Exception as e:
        print(f"  Warning: could not fetch T-bill rate — {e}")

    # Pre-classify non-market symbols (no API call needed)
    # Cash positions: beta = 0, price = $1.00, yield = current T-bill rate
    for s in symbols:
        if s in NON_MARKET_SYMBOLS:
            results[s]["ticker_type"]    = classify_ticker(s, {})
            results[s]["beta"]           = 0.0
            results[s]["live_price"]     = 1.0
            results[s]["dividend_yield"] = treasury_yield

    if not tradeable:
        return results

    # Build mapping: original symbol → Yahoo symbol (handles BRK.B → BRK-B etc.)
    yf_map = {s: _normalize_yf_symbol(s) for s in tradeable}
    yf_syms = list(yf_map.values())

    print(f"  Fetching market data for {len(tradeable)} tickers…")
    tickers_obj = yf.Tickers(" ".join(yf_syms))

    for orig_sym, yf_sym in yf_map.items():
        try:
            t = tickers_obj.tickers[yf_sym]
            info = t.info

            live_price = (
                info.get("currentPrice")
                or info.get("regularMarketPrice")
                or info.get("navPrice")       # mutual funds publish NAV
                or info.get("previousClose")
            )

            results[orig_sym]["beta"]           = _safe_beta(info, yf_sym, t)
            results[orig_sym]["live_price"]     = live_price
            results[orig_sym]["dividend_yield"] = _safe_dividend_yield(info)
            results[orig_sym]["ticker_type"]    = classify_ticker(orig_sym, info)

        except Exception as e:
            print(f"    Warning: could not fetch {orig_sym} ({yf_sym}) — {e}")

    return results


# ─── EXCEL ENRICHMENT ────────────────────────────────────────────────────────

def enrich_excel(path: str, market_data: dict, symbols_col_index: int = 2):
    """
    Adds columns I (Beta), J (Live Price), K (Dividend Yield %), L (Updated At)
    to an existing xlsx produced by the pipeline.
    symbols_col_index: 0-based index of the Symbol column in df (column B = index 1).
    """
    wb = load_workbook(path)
    ws = wb.active

    # ── Header styling helpers ────────────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    DATA_FONT    = Font(name="Arial", size=10)
    CENTER       = Alignment(horizontal="center", vertical="center")
    thin         = Side(style="thin", color="D9D9D9")
    BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)
    YELLOW_FILL  = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")

    new_headers = {
        9:  "Beta",
        10: "Live Price ($)",
        11: "Dividend Yield (%)",
        12: "Data Updated At",
        13: "Ticker Type",
        14: "Current Mkt Value ($)",   # live_price × quantity
        15: "Total Gain / Loss ($)",   # current_mkt_value − cost_basis
    }

    # Write headers row 1
    for col, label in new_headers.items():
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = 20

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Iterate data rows (row 2 onwards)
    for row_idx in range(2, ws.max_row + 1):
        sym_cell = ws.cell(row=row_idx, column=2)   # column B = Symbol
        sym = str(sym_cell.value).strip() if sym_cell.value else ""
        data = market_data.get(sym, {})

        beta  = data.get("beta")
        price = data.get("live_price")
        dyield = data.get("dividend_yield")

        # fetch_market_data already returns yield as a percentage float (e.g. 3.25)
        # No further multiplication needed here
        dyield_pct = round(dyield, 2) if dyield is not None else None

        ticker_type = data.get("ticker_type")

        # Current Market Value: if price is $1 (cash/MM) or missing, use Current Balance
        # as the authoritative value rather than computing 1 × quantity
        qty_cell    = ws.cell(row=row_idx, column=5)   # Quantity   = col E
        cost_cell   = ws.cell(row=row_idx, column=7)   # Cost Basis = col G
        bal_cell    = ws.cell(row=row_idx, column=8)   # Current Balance = col H
        qty         = qty_cell.value  if qty_cell.value  is not None else 0.0
        cost        = cost_cell.value if cost_cell.value is not None else 0.0
        cur_balance = bal_cell.value  if bal_cell.value  is not None else 0.0

        if price is None or price == 1.0:
            mkt_value = round(cur_balance, 2)
        else:
            mkt_value = round(price * qty, 2)
        total_gain = round(mkt_value - cost, 2) if mkt_value is not None else None

        values = {9: beta, 10: price, 11: dyield_pct, 12: timestamp,
                  13: ticker_type, 14: mkt_value, 15: total_gain}

        # Color map for ticker type column
        TYPE_COLORS = {
            "Stock":       "DDEEFF",   # light blue
            "ETF":         "D9EAD3",   # light green
            "Mutual Fund": "EAD1DC",   # light pink
            "Money Market":"FFF2CC",   # yellow
            "Cash":        "F3F3F3",   # light grey
            "Crypto":      "FCE5CD",   # light orange
            "Other":       "FFFFFF",   # white
        }
        for col, val in values.items():
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font   = DATA_FONT
            cell.border = BORDER
            cell.alignment = CENTER
            if val is None and col != 12:
                cell.fill = YELLOW_FILL   # highlight missing market data

            # Number formatting
            if col == 10 and val is not None:
                cell.number_format = '$#,##0.00'
            elif col == 11 and val is not None:
                cell.number_format = '0.00"%"'
            elif col == 9 and val is not None:
                cell.number_format = '0.00'
            elif col == 13 and val is not None:
                color = TYPE_COLORS.get(val, "FFFFFF")
                cell.fill = PatternFill("solid", start_color=color, end_color=color)
                cell.font = Font(name="Arial", size=10, bold=True)
            elif col == 14 and val is not None:
                cell.number_format = '$#,##0.00'
            elif col == 15 and val is not None:
                cell.number_format = '$#,##0.00;[RED]($#,##0.00)'
                # Bold positive gains green, losses red
                if val >= 0:
                    cell.font = Font(name="Arial", size=10, color="207020")
                else:
                    cell.font = Font(name="Arial", size=10, color="CC0000")

    wb.save(path)
    print(f"  Excel enriched → {path}")


# ─── SUMMARY SHEET ───────────────────────────────────────────────────────────

def build_summary_sheet(path, df_positions, market_data, april_total):
    """
    Adds a 'Portfolio Summary' sheet with:
      - Total Current Market Value  (live_price × quantity, summed)
      - April snapshot value        (from hardcoded total_net_worth)
      - Total Gain / Loss           (current − april)
      - % Change
      - Breakdown by Ticker Type
    """
    wb = load_workbook(path)

    # Remove existing summary sheet if re-running
    if "Portfolio Summary" in wb.sheetnames:
        del wb["Portfolio Summary"]
    ws = wb.create_sheet("Portfolio Summary", 0)  # insert as first tab

    # ── Styles ──────────────────────────────────────────────────────────────
    HDR_FILL   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    TITLE_FONT = Font(bold=True, name="Arial", size=14)
    LABEL_FONT = Font(bold=True, name="Arial", size=11)
    DATA_FONT  = Font(name="Arial", size=11)
    GREEN_FONT = Font(name="Arial", size=11, color="207020", bold=True)
    RED_FONT   = Font(name="Arial", size=11, color="CC0000", bold=True)
    CENTER     = Alignment(horizontal="center", vertical="center")
    LEFT       = Alignment(horizontal="left",   vertical="center")
    thin       = Side(style="thin", color="D9D9D9")
    BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)
    CURR_FMT   = '$#,##0.00'
    PCT_FMT    = '+0.00"%";-0.00"%"'

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22

    # ── Title ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value     = "Portfolio Summary"
    title_cell.font      = TITLE_FONT
    title_cell.alignment = CENTER
    title_cell.fill      = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")

    ws.merge_cells("A2:C2")
    ts_cell = ws["A2"]
    ts_cell.value     = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ts_cell.font      = Font(name="Arial", size=9, italic=True)
    ts_cell.alignment = CENTER
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    # ── Compute current market value per row ─────────────────────────────────
    total_current = 0.0
    type_totals   = {}   # ticker_type → current mkt value

    for _, row in df_positions.iterrows():
        sym  = str(row["Symbol"]).strip() if row["Symbol"] else ""
        qty  = row["Quantity"] or 0.0
        cost = row["Cost Basis"] or 0.0
        data = market_data.get(sym, {})
        price = data.get("live_price")
        ttype = data.get("ticker_type") or "Other"

        if price is None or price == 1.0:
            mkt_val = row["Current Balance"] or 0.0
        else:
            mkt_val = price * qty
        total_current += mkt_val
        type_totals[ttype] = type_totals.get(ttype, 0.0) + mkt_val

    total_gain   = total_current - april_total
    pct_change   = (total_gain / april_total * 100) if april_total else 0.0

    # ── Top-level KPIs (rows 4–8) ────────────────────────────────────────────
    kpis = [
        ("April Snapshot Value",  april_total,   CURR_FMT, DATA_FONT),
        ("Current Market Value",  total_current, CURR_FMT, DATA_FONT),
        ("Total Gain / Loss ($)", total_gain,    '$#,##0.00;[RED]($#,##0.00)',
             GREEN_FONT if total_gain >= 0 else RED_FONT),
        ("Total Gain / Loss (%)", pct_change,    PCT_FMT,
             GREEN_FONT if pct_change >= 0 else RED_FONT),
    ]

    row_n = 4
    # Section header
    for c, label in enumerate(["Metric", "Value"], start=1):
        hc = ws.cell(row=row_n, column=c, value=label)
        hc.fill = HDR_FILL; hc.font = HDR_FONT; hc.alignment = CENTER; hc.border = BORDER
    ws.merge_cells(f"C{row_n}:C{row_n}")
    row_n += 1

    for label, value, fmt, font in kpis:
        lc = ws.cell(row=row_n, column=1, value=label)
        lc.font = LABEL_FONT; lc.alignment = LEFT; lc.border = BORDER

        vc = ws.cell(row=row_n, column=2, value=value)
        vc.font = font; vc.alignment = CENTER
        vc.number_format = fmt; vc.border = BORDER

        ws.row_dimensions[row_n].height = 20
        row_n += 1

    # ── Breakdown by Ticker Type (rows row_n+1 onwards) ──────────────────────
    row_n += 1
    for c, label in enumerate(["Ticker Type", "Market Value ($)", "% of Portfolio"], start=1):
        hc = ws.cell(row=row_n, column=c, value=label)
        hc.fill = HDR_FILL; hc.font = HDR_FONT; hc.alignment = CENTER; hc.border = BORDER
    row_n += 1

    TYPE_COLORS = {
        "Stock":        "DDEEFF",
        "ETF":          "D9EAD3",
        "Mutual Fund":  "EAD1DC",
        "Money Market": "FFF2CC",
        "Cash":         "F3F3F3",
        "Crypto":       "FCE5CD",
        "Other":        "FFFFFF",
    }
    TYPE_ORDER = ["Stock", "ETF", "Mutual Fund", "Money Market", "Cash", "Crypto", "Other"]

    for ttype in TYPE_ORDER:
        val = type_totals.get(ttype, 0.0)
        if val == 0.0:
            continue
        pct = val / total_current * 100 if total_current else 0.0
        color = TYPE_COLORS.get(ttype, "FFFFFF")
        fill  = PatternFill("solid", start_color=color, end_color=color)

        tc = ws.cell(row=row_n, column=1, value=ttype)
        tc.font = Font(name="Arial", size=11, bold=True)
        tc.fill = fill; tc.alignment = LEFT; tc.border = BORDER

        vc = ws.cell(row=row_n, column=2, value=val)
        vc.number_format = CURR_FMT
        vc.font = DATA_FONT; vc.fill = fill; vc.alignment = CENTER; vc.border = BORDER

        pc = ws.cell(row=row_n, column=3, value=pct)
        pc.number_format = '0.0"%"'
        pc.font = DATA_FONT; pc.fill = fill; pc.alignment = CENTER; pc.border = BORDER

        ws.row_dimensions[row_n].height = 20
        row_n += 1

    # Total row
    tc = ws.cell(row=row_n, column=1, value="TOTAL")
    tc.font = Font(name="Arial", size=11, bold=True); tc.alignment = LEFT; tc.border = BORDER
    vc = ws.cell(row=row_n, column=2, value=total_current)
    vc.number_format = CURR_FMT
    vc.font = Font(name="Arial", size=11, bold=True); vc.alignment = CENTER; vc.border = BORDER
    pc = ws.cell(row=row_n, column=3, value=100.0)
    pc.number_format = '0.0"%"'
    pc.font = Font(name="Arial", size=11, bold=True); pc.alignment = CENTER; pc.border = BORDER

    wb.save(path)
    print(f"  Summary sheet written → {path}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    db_conn = setup_database()

    file_vanguard = "04182026 Holdings | Vanguard.pdf"
    file_benefits = "04222026 Benefits OnLine | Investments.pdf"
    file_fidelity = "04272026 Fidelity Portfolio Positions.pdf"

    text_vanguard = read_pdf_text(file_vanguard)
    text_benefits = read_pdf_text(file_benefits)
    text_fidelity = read_pdf_text(file_fidelity)

    parse_vanguard(db_conn, text_vanguard)
    parse_benefits_online(db_conn, text_benefits)
    parse_fidelity(db_conn, text_fidelity)

    print("\nBuilding unified positions matrix…")
    query = """
        SELECT
            a.account_name          AS [Account Name],
            h.symbol                AS [Symbol],
            h.name                  AS [Asset Name],
            h.price                 AS [Price],
            h.quantity              AS [Quantity],
            h.unrealized_gain_loss  AS [$ Unrealized gain/loss],
            h.cost_basis            AS [Cost Basis],
            h.current_balance       AS [Current Balance]
        FROM holdings h
        JOIN accounts a ON h.account_id = a.account_id
        ORDER BY a.total_balance DESC, h.current_balance DESC
    """
    df_final = pd.read_sql_query(query, db_conn)
    output_path = "output.xlsx"
    df_final.to_excel(output_path, index=False)

    # ── Market data enrichment ────────────────────────────────────────────
    print("\nFetching live market data (Beta, Price, Yield)…")
    unique_symbols = df_final["Symbol"].dropna().unique().tolist()
    market_data = fetch_market_data(unique_symbols)
    enrich_excel(output_path, market_data)

    # ── Portfolio summary sheet ───────────────────────────────────────────
    print("\nBuilding portfolio summary sheet…")
    april_total = 5712900.84   # total_net_worth as of 2026-04-18
    build_summary_sheet(output_path, df_final, market_data, april_total)

    print("-" * 60)
    print("FINISHED SUCCESSFULLY!")
    print(f"Output: '{output_path}'")
    print(f"  Sheet 1 : Portfolio Summary (KPIs + breakdown by type)")
    print(f"  Sheet 2 : Holdings (columns A–O)")
    print(f"    A–H : Portfolio positions (from brokerage data)")
    print(f"    I   : Beta")
    print(f"    J   : Live Price ($)")
    print(f"    K   : Dividend Yield (%)")
    print(f"    L   : Data timestamp")
    print(f"    M   : Ticker Type")
    print(f"    N   : Current Mkt Value  (live price × quantity)")
    print(f"    O   : Total Gain / Loss  (current mkt value − cost basis)")
    print("-" * 60)

    db_conn.close()
    